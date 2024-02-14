# pip install pandas seatable-api
# pip install python-dotenv
import json
from seatable_api import Base
from openpyxl import load_workbook
import pandas as pd
from dotenv import load_dotenv
import os

# 加载 .env 文件中的环境变量
load_dotenv()

# 从 .env 文件中读取敏感配置
mysql_config = {
    'host': os.getenv('MYSQL_HOST'),
    'port': int(os.getenv('MYSQL_PORT')),  # 确保转换为正确的类型
    'user': os.getenv('MYSQL_USER'),
    'password': os.getenv('MYSQL_PASSWORD'),
    'db': os.getenv('MYSQL_DB')  # 假设你在 .env 中也定义了 MYSQL_DB
}

seatable_config = {
    'server_url': os.getenv('SEATABLE_SERVER_URL'),
    'api_token': os.getenv('SEATABLE_API_TOKEN')
}

# 读取配置文件
print("Loading configuration file...")
with open('config.json', 'r') as f:
    other_config = json.load(f)

# 从配置文件获取配置信息
#server_url = config['SERVER_URL']
#api_token = config['TEST_TOKEN']
table_name = other_config['table_name']
excel_file_path = other_config['excel_file_path']
sheet_name = other_config['sheet_name']
relation_field = other_config['relation_field']
field_mappings = other_config['field_mappings']


def sync_xlsx():
    """Sync table into the xlsx
    """
    # base initiated and authed
    print(f"Connecting to Seatable: {seatable_config['server_url']}...")
    base = Base(seatable_config['api_token'], seatable_config['server_url'])
    base.auth()

    # 从Seatable获取数据
    print("Fetching data from Seatable...")
    rows = base.list_rows(table_name)
    seatable_df = pd.DataFrame(rows)
    seatable_df[relation_field] = seatable_df[relation_field].astype(str)  # 确保关联字段为字符串

    # 读取Excel文件
    print("Loading Excel file...")
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook[sheet_name]

    # 更新Excel
    print("Starting to update Excel...")
    updated_rows_count = 0  # Count of updated rows
    for seatable_row in seatable_df.itertuples(index=False):
        # 通过关联字段找到对应的Excel行
        for excel_row in range(2, sheet.max_row + 1):
            excel_relation_value = str(sheet.cell(row=excel_row, column=field_mappings[relation_field]).value)
            if getattr(seatable_row, relation_field) == excel_relation_value:
                # 根据配置更新字段
                for seatable_field, excel_col in field_mappings.items():
                    excel_value = getattr(seatable_row, seatable_field, None)
                    if excel_value is not None:
                        sheet.cell(row=excel_row, column=excel_col, value=excel_value)
                updated_rows_count += 1


    # 保存Excel文件
    print(f"Saving changes to the Excel file: {excel_file_path}...")
    workbook.save(filename=excel_file_path)

    print(f"Completed! Total of {updated_rows_count} rows were updated.")


if __name__ == '__main__':
    sync_xlsx()
